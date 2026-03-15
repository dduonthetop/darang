from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[3]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from darang.faq.admin_store import CSV_HEADERS, _write_static_data  # noqa: E402
from darang.faq.faq_loader import load_faq_items  # noqa: E402


FAQ_DIR = REPO_ROOT / "darang" / "faq"
CSV_PATH = FAQ_DIR / "ipark_faq_dataset.csv"
ADMIN_XLSX_PATH = FAQ_DIR / "ipark_faq_dataset_admin_v2.xlsx"
RAW_MANUAL_BASE = "https://raw.githubusercontent.com/dduonthetop/darang/main/darang/faq/references/manual/"
PHONE_MANUAL = RAW_MANUAL_BASE + "%EB%82%B4%EC%84%A0%20%EC%A0%84%ED%99%94%20%EC%8B%A0%EC%B2%AD%20%EB%B0%A9%EB%B2%95.pdf"
POS_MANUAL = RAW_MANUAL_BASE + "%EC%95%84%EC%9D%B4%ED%8C%8C%ED%81%AC%EB%AA%B0%20%EC%9A%A9%EC%82%B0%EC%A0%90%20%EB%A6%AC%EB%B9%99%2C%ED%8C%A8%EC%85%98%2C%ED%8C%9D%EC%97%85-POS%20%EB%A7%A4%EB%89%B4%EC%96%BC.pdf"
FNB_POS_MANUAL = RAW_MANUAL_BASE + "%EC%95%84%EC%9D%B4%ED%8C%8C%ED%81%AC%EB%AA%B0%20%EC%9A%A9%EC%82%B0%EC%A0%90%20FnB-POS%20%EB%A7%A4%EB%89%B4%EC%96%BC_%ED%98%91%EB%A0%A5%EC%82%AC%EC%9B%90%20%EA%B5%90%EC%9C%A1%EC%9A%A9.pdf"
MANUAL_REWRITES = {
    "./references/manual/내선 전화 신청 방법.pdf": PHONE_MANUAL,
    "./references/manual/아이파크몰 용산점 리빙,패션,팝업-POS 매뉴얼.pdf": POS_MANUAL,
    "./references/manual/아이파크몰 용산점 FnB-POS 매뉴얼_협력사원 교육용.pdf": FNB_POS_MANUAL,
}

BASE_COLUMNS = [
    "faq_id",
    "stage",
    "audience",
    "category_code",
    "category_name",
    "question",
    "paraphrase_1",
    "paraphrase_2",
    "paraphrase_3",
    "paraphrase_4",
    "paraphrase_5",
    "paraphrase_6",
    "paraphrase_7",
    "paraphrase_8",
    "paraphrases_full",
    "answer",
    "next_action",
    "contact_channel",
    "restrictions",
    "visibility",
    "confidence_type",
    "review_status",
    "needs_revision",
    "last_reviewer",
    "notes",
    "updated_at",
    "source",
]

WIDTH_MAP = {
    "faq_id": 12,
    "stage": 12,
    "audience": 10,
    "category_code": 12,
    "category_name": 24,
    "question": 36,
    "paraphrase_1": 28,
    "paraphrase_2": 28,
    "paraphrase_3": 28,
    "paraphrase_4": 28,
    "paraphrase_5": 28,
    "paraphrase_6": 28,
    "paraphrase_7": 28,
    "paraphrase_8": 28,
    "paraphrases_full": 52,
    "answer": 54,
    "next_action": 42,
    "contact_channel": 28,
    "restrictions": 24,
    "visibility": 14,
    "confidence_type": 20,
    "review_status": 14,
    "needs_revision": 12,
    "last_reviewer": 14,
    "notes": 28,
    "updated_at": 14,
    "source": 28,
}


def normalize_question_text(text: str) -> str:
    return " ".join((text or "").strip().replace("\n", " ").split())


def has_batchim(text: str) -> bool:
    text = text.strip()
    if not text:
        return False
    last = text[-1]
    if not ("가" <= last <= "힣"):
        return False
    return (ord(last) - ord("가")) % 28 != 0


def topic_particle(text: str) -> str:
    return "은" if has_batchim(text) else "는"


def object_particle(text: str) -> str:
    return "을" if has_batchim(text) else "를"


def build_question_stem(question: str) -> str:
    stem = normalize_question_text(question).rstrip("?").rstrip(".")
    patterns = [
        r"(은|는|이|가)?\s*각각\s*어떻게\s*(되나요|되는가|되는지)$",
        r"(은|는|이|가)?\s*어떻게\s*(되나요|되는가|진행되나요|안내되나요|확인하나요)$",
        r"(은|는|이|가)?\s*정해져\s*있나요$",
        r"(은|는|이|가)?\s*가능한가요$",
        r"(은|는|이|가)?\s*필요한가요$",
        r"(은|는|이|가)?\s*있나요$",
        r"(은|는|이|가)?\s*인가요$",
        r"(은|는|이|가)?\s*무엇인가요$",
        r"(은|는|이|가)?\s*어디인가요$",
        r"(은|는|이|가)?\s*언제인가요$",
        r"(은|는|이|가)?\s*알려주세요$",
        r"(은|는|이|가)?\s*문의드립니다$",
    ]
    for pattern in patterns:
        next_stem = re.sub(pattern, "", stem).strip(" ,")
        if next_stem and next_stem != stem:
            stem = next_stem
            break
    return stem.strip(" ,")


def generate_paraphrases(question: str) -> List[str]:
    question = normalize_question_text(question)
    stem = build_question_stem(question)
    topic = topic_particle(stem)
    obj = object_particle(stem)
    summary_text = f"{stem} 내용이 궁금합니다." if stem.endswith("기준") else f"{stem} 관련 기준이 궁금합니다."
    seeds = [
        question,
        f"{stem}{topic} 어떻게 되나요?",
        f"{stem}{obj} 알려주세요.",
        f"{stem} 문의드립니다.",
        f"{stem} 확인 부탁드립니다.",
        f"{stem}{topic} 어디에서 확인하나요?",
        f"{stem}{obj} 안내해 주세요.",
        f"{stem} 알 수 있을까요?",
        summary_text,
        f"{stem} 관련해서 확인하고 싶습니다.",
    ]
    unique: List[str] = []
    seen = set()
    for item in seeds:
        normalized = normalize_question_text(item)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        unique.append(normalized)
    return unique[:10]


def canonical_category(question: str, category: str) -> str:
    question_norm = normalize_question_text(question)
    if "POS" in question_norm or "포스" in question_norm:
        return "8.12 POS"
    if "내선전화" in question_norm or "내선 전화" in question_norm or "KT" in question_norm:
        return "8.13 유선전화"
    return category


def canonical_manual_files(question: str, existing_manuals: str) -> str:
    if existing_manuals.strip():
        pieces = [piece.strip() for piece in existing_manuals.split(";") if piece.strip()]
        rewritten = [MANUAL_REWRITES.get(piece, piece) for piece in pieces]
        return "; ".join(rewritten)
    question_norm = normalize_question_text(question)
    if "POS" in question_norm or "포스" in question_norm:
        return "; ".join([POS_MANUAL, FNB_POS_MANUAL])
    if "내선전화" in question_norm or "내선 전화" in question_norm or "KT" in question_norm:
        return PHONE_MANUAL
    return existing_manuals.strip()


def load_csv_rows() -> Dict[str, Dict[str, str]]:
    with CSV_PATH.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return {str(row.get("faq_id", "")).strip(): row for row in reader}


def load_admin_rows() -> List[Dict[str, str]]:
    workbook = load_workbook(ADMIN_XLSX_PATH, data_only=True)
    worksheet = workbook["FAQ_Edit"]
    headers = [cell.value for cell in worksheet[1]]
    rows: List[Dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        row = {str(headers[idx]): ("" if values[idx] is None else str(values[idx]).strip()) for idx in range(len(headers))}
        if not row.get("faq_id"):
            continue
        rows.append(row)
    return rows


def build_csv_rows(admin_rows: List[Dict[str, str]], csv_rows: Dict[str, Dict[str, str]]) -> List[Dict[str, str]]:
    merged_rows: List[Dict[str, str]] = []
    for row in admin_rows:
        faq_id = row["faq_id"]
        existing = dict(csv_rows.get(faq_id, {}))
        category_code = row.get("category_code", "").strip()
        category_name = row.get("category_name", "").strip()
        category = f"{category_code} {category_name}".strip()
        category = canonical_category(row.get("question", ""), category)
        paraphrases = generate_paraphrases(row.get("question", ""))
        merged = {key: existing.get(key, "") for key in CSV_HEADERS}
        merged.update(
            {
                "faq_id": faq_id,
                "stage": row.get("stage", "").strip(),
                "audience": row.get("audience", "").strip(),
                "category": category,
                "question": normalize_question_text(row.get("question", "")),
                "paraphrases": "; ".join(paraphrases),
                "answer": normalize_question_text(row.get("answer", "")),
                "next_action": normalize_question_text(row.get("next_action", "")),
                "contact_channel": normalize_question_text(row.get("contact_channel", "")),
                "restrictions": normalize_question_text(row.get("restrictions", "")),
                "visibility": normalize_question_text(row.get("visibility", "")),
                "confidence_type": normalize_question_text(row.get("confidence_type", "")),
                "updated_at": normalize_question_text(row.get("updated_at", "")) or existing.get("updated_at", ""),
                "source": normalize_question_text(row.get("source", "")) or existing.get("source", ""),
                "manual_files": canonical_manual_files(row.get("question", ""), existing.get("manual_files", "")),
            }
        )
        merged_rows.append(merged)
    return merged_rows


def write_csv(rows: List[Dict[str, str]]) -> None:
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_admin_workbook(rows: List[Dict[str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FAQ_Edit"
    guide = workbook.create_sheet("Guide")
    lookup = workbook.create_sheet("Lookup")

    header_fill = PatternFill(fill_type="solid", fgColor="DCE8F5")
    soft_fill = PatternFill(fill_type="solid", fgColor="F8FBFF")
    important_fill = PatternFill(fill_type="solid", fgColor="FFF1EE")
    header_font = Font(bold=True, color="1F2937")
    small_font = Font(size=10, color="4B5563")
    thin = Side(style="thin", color="D7DCE2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(vertical="center", horizontal="center", wrap_text=True)

    worksheet.append(BASE_COLUMNS)
    for col_idx, name in enumerate(BASE_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = important_fill if name in {"review_status", "needs_revision", "last_reviewer", "notes"} else header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_index, row in enumerate(rows, start=2):
        category = row.get("category", "").strip()
        if " " in category:
            category_code, category_name = category.split(" ", 1)
        else:
            category_code, category_name = category, ""
        paraphrases = [piece.strip() for piece in row.get("paraphrases", "").split(";") if piece.strip()]
        admin_row = {
            "faq_id": row.get("faq_id", ""),
            "stage": row.get("stage", ""),
            "audience": row.get("audience", ""),
            "category_code": category_code,
            "category_name": category_name,
            "question": row.get("question", ""),
            "paraphrase_1": paraphrases[0] if len(paraphrases) > 0 else "",
            "paraphrase_2": paraphrases[1] if len(paraphrases) > 1 else "",
            "paraphrase_3": paraphrases[2] if len(paraphrases) > 2 else "",
            "paraphrase_4": paraphrases[3] if len(paraphrases) > 3 else "",
            "paraphrase_5": paraphrases[4] if len(paraphrases) > 4 else "",
            "paraphrase_6": paraphrases[5] if len(paraphrases) > 5 else "",
            "paraphrase_7": paraphrases[6] if len(paraphrases) > 6 else "",
            "paraphrase_8": paraphrases[7] if len(paraphrases) > 7 else "",
            "paraphrases_full": "; ".join(paraphrases),
            "answer": row.get("answer", ""),
            "next_action": row.get("next_action", ""),
            "contact_channel": row.get("contact_channel", ""),
            "restrictions": row.get("restrictions", ""),
            "visibility": row.get("visibility", ""),
            "confidence_type": row.get("confidence_type", ""),
            "review_status": "검토 전",
            "needs_revision": "",
            "last_reviewer": "",
            "notes": "",
            "updated_at": row.get("updated_at", ""),
            "source": row.get("source", ""),
        }
        worksheet.append([admin_row[column] for column in BASE_COLUMNS])
        for cell in worksheet[row_index]:
            cell.border = border
            cell.alignment = wrap
            if row_index % 2 == 0:
                cell.fill = soft_fill

    for idx, name in enumerate(BASE_COLUMNS, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = WIDTH_MAP.get(name, 18)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.zoomScale = 88
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 42

    lookup["A1"] = "confidence_type"
    lookup["A2"] = "운영 사례형 답변"
    lookup["A3"] = "확정 정책형 답변"
    lookup["A4"] = "확인 필요형 답변"
    lookup["B1"] = "review_status"
    lookup["B2"] = "검토 전"
    lookup["B3"] = "검토 중"
    lookup["B4"] = "검토 완료"
    lookup["B5"] = "수정 필요"
    lookup["C1"] = "needs_revision"
    lookup["C2"] = "Y"
    lookup["C3"] = "N"
    lookup.sheet_state = "hidden"

    confidence_validation = DataValidation(type="list", formula1="=Lookup!$A$2:$A$4", allow_blank=True)
    review_validation = DataValidation(type="list", formula1="=Lookup!$B$2:$B$5", allow_blank=True)
    revision_validation = DataValidation(type="list", formula1="=Lookup!$C$2:$C$3", allow_blank=True)
    worksheet.add_data_validation(confidence_validation)
    worksheet.add_data_validation(review_validation)
    worksheet.add_data_validation(revision_validation)
    confidence_validation.add(f"U2:U{worksheet.max_row}")
    review_validation.add(f"V2:V{worksheet.max_row}")
    revision_validation.add(f"W2:W{worksheet.max_row}")

    for row in [
        ["관리자용 FAQ 편집 가이드"],
        ["1. question, answer, next_action을 수정한 뒤 이 스크립트를 다시 실행하면 paraphrase가 자동 갱신됩니다."],
        ["2. paraphrase_1~8은 자동 생성 결과를 보여주는 영역이며, 전체 문자열은 paraphrases_full에서 확인할 수 있습니다."],
        ["3. review_status, needs_revision, notes는 검토 협업용입니다."],
        ["4. 이 파일은 편집용 원본이며 실제 서비스 반영은 CSV와 local_faq_data.js로 동기화됩니다."],
    ]:
        guide.append(row)
    for cell in guide["A"]:
        cell.alignment = wrap
        if cell.row == 1:
            cell.font = Font(bold=True, size=14, color="1F2937")
            cell.fill = header_fill
        else:
            cell.font = small_font
    guide.column_dimensions["A"].width = 95

    workbook.save(ADMIN_XLSX_PATH)


def main() -> None:
    csv_rows = load_csv_rows()
    admin_rows = load_admin_rows()
    merged_rows = build_csv_rows(admin_rows, csv_rows)
    write_csv(merged_rows)
    write_admin_workbook(merged_rows)
    _write_static_data(load_faq_items(CSV_PATH))
    print(f"Synced {len(merged_rows)} FAQ rows from {ADMIN_XLSX_PATH.name}")


if __name__ == "__main__":
    main()
